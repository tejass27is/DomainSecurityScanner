"""
VAPT report parsing engine.

Supported input formats:
- Nessus v2 XML (``.nessus`` / ``.xml``) — parsed XXE-safe via ``defusedxml``
  (malicious XML with external entities / entity expansion is rejected).
- CSV (``.csv``) — flexible header detection auto-detects Nessus / OpenVAS /
  Qualys / generic exports.
- Excel (``.xlsx``) — flexible column detection via ``openpyxl``.

Every parser returns a list of *raw* findings with a common shape::

    {
        "host": str, "os": str, "port": int|None, "protocol": str,
        "service": str, "plugin_id": str, "plugin_name": str,
        "plugin_family": str, "title": str, "severity_label": str,
        "severity": int (0-4), "cvss_score": float|None, "cvss_vector": str,
        "description": str, "synopsis": str, "solution": str,
        "references": [str], "cves": [str], "evidence": str,
    }
"""

import csv
import io
import os
import re
from pathlib import Path

try:
    from defusedxml import ElementTree as SafeET
except ImportError:
    import xml.etree.ElementTree as SafeET

# Max upload size in MB (configurable via VAPT_MAX_FILE_SIZE_MB).
VAPT_MAX_FILE_SIZE_MB = int(os.getenv("VAPT_MAX_FILE_SIZE_MB", "25"))
MAX_FILE_SIZE = VAPT_MAX_FILE_SIZE_MB * 1024 * 1024
ALLOWED_EXTENSIONS = {".nessus", ".xml", ".csv", ".xlsx", ".xls"}

SEVERITY_LABELS = {0: "info", 1: "low", 2: "medium", 3: "high", 4: "critical"}
SEVERITY_RANK = {"info": 0, "low": 1, "medium": 2, "high": 3, "critical": 4}
SOURCE_TOOL_LABELS = {
    "nessus": "Nessus",
    "openvas": "OpenVAS",
    "qualys": "Qualys",
    "generic": "Generic",
}


# ─── Small helpers ────────────────────────────────────────────────────────────

def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _int_or_none(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _split_refs(value) -> list[str]:
    """Split a 'see_also' style block into individual reference URLs."""
    if not value:
        return []
    parts = re.split(r"[\s,;]+", _clean_text(value))
    return [p for p in parts if p]


# ─── Severity mapping ─────────────────────────────────────────────────────────

def severity_from_cvss(cvss_score) -> tuple[str, int]:
    """Map a CVSS base score to (label, 0-4) using CVSS v3 boundaries."""
    score = _float(cvss_score)
    if score is None or score <= 0:
        return ("info", 0)
    if score < 4.0:
        return ("low", 1)
    if score < 7.0:
        return ("medium", 2)
    if score < 9.0:
        return ("high", 3)
    return ("critical", 4)


def parse_severity(value, cvss_score=None) -> tuple[str, int]:
    """
    Coerce an arbitrary severity value into a (label, 0-4) tuple.

    Accepts Nessus 0-4 integers, plain labels (Critical / High / Medium / Low /
    Info), and falls back to a CVSS-derived severity when a score is available.
    """
    if value is None or str(value).strip() == "":
        if cvss_score is not None:
            return severity_from_cvss(cvss_score)
        return ("info", 0)

    text = str(value).strip().lower()

    # Numeric forms
    if re.fullmatch(r"\d+(\.\d+)?", text):
        num = float(text)
        if 0 <= num <= 4 and num.is_integer():
            # Nessus-style 0-4 severity
            return (SEVERITY_LABELS[int(num)], int(num))
        mapped = severity_from_cvss(num)
        if mapped:
            return mapped

    # Label forms (longest first so 'high' isn't swallowed by 'critical')
    for label, rank in (("critical", 4), ("high", 3), ("medium", 2), ("moderate", 2),
                        ("low", 1), ("info", 0), ("informational", 0), ("none", 0),
                        ("log", 0), ("debug", 0)):
        if label in text:
            return (SEVERITY_LABELS[rank], rank)

    if cvss_score is not None:
        return severity_from_cvss(cvss_score)
    return ("info", 0)


# ─── Nessus v2 XML ────────────────────────────────────────────────────────────

def _host_os(host_el) -> str:
    """Best-effort OS detection from a Nessus ``ReportHost`` element.

    Nessus stores host metadata as ``<tag name="..." value="..."/>`` children
    (``operating-system``, ``os``, ``Host_OS``, ``os.name``...). The
    ``operating-system`` tag is the canonical source; others are used as a
    fallback.
    """
    fallback = ""
    for tag_el in host_el.findall(".//tag"):
        name = (tag_el.get("name") or "").strip().lower()
        value = (tag_el.get("value") or tag_el.text or "").strip()
        if not value:
            continue
        if name == "operating-system":
            return value
        if not fallback and name in ("os", "host_os", "os.name", "os.product",
                                     "os_product", "operating system"):
            fallback = value
    return fallback


def _strip_tag_name(tag_name: str) -> str:
    if not tag_name:
        return ""
    return tag_name.split("}")[-1].strip().lower()


def _tag_text(tag_el) -> str:
    if tag_el is None:
        return ""
    return (tag_el.get("value") or tag_el.text or "").strip()


def _tag_name(tag_el) -> str:
    name = tag_el.get("name") or tag_el.findtext("name") or ""
    return str(name).strip().lower()


def _host_mac(host_el) -> str:
    """Extract Nessus host MAC address metadata from a ReportHost element."""
    allowed_keys = {"mac address", "mac", "macaddress", "mac addr", "mac_addr", "mac-address"}
    candidate_tags = {"tag", "property"}

    for element in host_el.iter():
        if _strip_tag_name(element.tag) not in candidate_tags:
            continue
        name = _tag_name(element)
        value = _tag_text(element)
        if not value:
            continue
        if name in allowed_keys or "mac" in name:
            return value

    return ""


def parse_nessus_xml(content: bytes) -> list[dict]:
    """Parse a Nessus v2 XML export (`.nessus` / `.xml`) into raw findings."""
    try:
        root = SafeET.fromstring(content)
    except Exception as exc:  # defusedxml raises on malicious XML
        raise ValueError(f"Invalid or unsafe XML in export file: {exc}")

    if root.tag != "NessusClientData_v2":
        raise ValueError(
            "This XML is not a Nessus v2 export (expected <NessusClientData_v2>)."
        )

    raw: list[dict] = []
    for host_el in root.iter("ReportHost"):
        host = (host_el.get("name") or "").strip()
        os_name = _host_os(host_el)
        for item in host_el.iter("ReportItem"):
            cvss = _float(item.findtext("cvss_base_score"))
            label, rank = parse_severity(item.get("severity", "0"), cvss)
            cves = [
                c.text.strip()
                for c in item.findall("cve")
                if c.text and c.text.strip()
            ]
            plugin_name = (item.get("pluginName") or "").strip()
            raw.append({
                "host": host,
                "os": os_name,
                "port": _int_or_none(item.get("port")),
                "protocol": (item.get("protocol") or "").strip(),
                "service": (item.get("svc_name") or "").strip(),
                "plugin_id": (item.get("pluginID") or "").strip(),
                "plugin_name": plugin_name,
                "plugin_family": (item.get("pluginFamily") or "").strip(),
                "title": plugin_name or _clean_text(item.findtext("plugin_name")),
                "severity_label": label,
                "severity": rank,
                "cvss_score": cvss,
                "cvss_vector": _clean_text(
                    item.findtext("cvss_vector") or item.findtext("cvss3_vector")
                ),
                "description": _clean_text(item.findtext("description")),
                "synopsis": _clean_text(item.findtext("synopsis")),
                "solution": _clean_text(item.findtext("solution")),
                "references": _split_refs(item.findtext("see_also")),
                "cves": cves,
                "mac_address": _host_mac(host_el),
                "evidence": _clean_text(item.findtext("plugin_output")),
            })
    return raw


# ─── Flexible header detection (CSV / XLSX) ───────────────────────────────────

# Canonical key -> list of accepted header aliases (matched as lowercase slugs,
# e.g. "Plugin ID" -> "plugin id").
HEADER_ALIASES: dict[str, list[str]] = {
    "host": ["host", "hostname", "ip", "ip address", "address", "asset",
             "dns name", "fqdn", "host ip", "host name", "ip address (dns name)"],
    "port": ["port", "port number"],
    "protocol": ["protocol", "proto"],
    "service": ["service", "svc", "svc name", "application protocol",
                "service name"],
    "plugin_id": ["plugin id", "vulnerability id", "qid", "oid", "nvt oid",
                  "vuln id", "check id", "pluginid", "plugin-id"],
    "title": ["plugin name", "name", "title", "finding", "vulnerability",
              "issue", "vulnerability name", "plugin", "check name",
              "vuln name", "nvt name", "test name", "vuln title", "plugin title"],
    "severity": ["severity", "risk", "risk factor", "level", "threat",
                 "risk level", "severity level", "risk score", "priority"],
    "cvss": ["cvss", "cvss base score", "cvss score", "base score", "score",
             "cvss v2", "cvss v3", "cvss2 base score", "cvss3 base score",
             "cvss v2 base score", "cvss v3 base score", "cvssv2", "cvssv3",
             "cvss2 score", "cvss3 score", "cvss temporal"],
    "cvss_vector": ["cvss vector", "cvss2 vector", "cvss3 vector",
                    "cvss v2 vector", "cvss v3 vector"],
    "description": ["description", "synopsis", "detail", "details",
                    "vulnerability details", "plugin description", "summary",
                    "impact", "vulnerability description", "description/synopsis"],
    "solution": ["solution", "fix", "remediation", "recommendation",
                 "how to fix", "remediation steps", "solution text",
                 "solution / workaround"],
    "references": ["references", "see also", "see_also", "reference",
                   "links", "reference links", "references (see also)"],
    "cves": ["cve", "cves", "cve ids", "cve id"],
    "evidence": ["plugin output", "output", "evidence", "proof",
                 "result", "proof of concept", "poc", "actual result",
                 "plugin_output", "output data"],
    "plugin_family": ["plugin family", "family", "vulnerability family",
                      "group"],
    "os": ["os", "operating system", "os name", "os version", "platform",
           "operating system (os)", "host os", "os family"],
    "status": ["status", "finding status", "state", "resolution"],
    "mac_address": ["mac address", "mac", "macaddress", "mac addr"],
    "hostname": ["hostname", "host name", "device name"],
    "operating_system": ["operating system", "os", "os name", "os version",
                         "platform", "host os", "operating system (os)", "os family"],
    "comment": ["remarks", "remark", "comment", "comments", "notes", "note"],
}


def _slug(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def detect_source_tool(headers: list[str]) -> str:
    """Best-effort source-tool detection from a CSV/XLSX header row."""
    slugs = {_slug(h) for h in headers if h}
    joined = " ".join(slugs)
    if "qid" in slugs or "qualys" in joined:
        return "qualys"
    if "nvt" in joined or "oid" in slugs:
        return "openvas"
    if "plugin id" in joined and ("host" in joined or "ip" in joined):
        return "nessus"
    return "generic"


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Map canonical keys -> column index, resolving header aliases.

    Column order determines priority; each canonical key is assigned at most
    once (first matching header wins).
    """
    taken = set()
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        slug = _slug(header)
        if not slug:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if key in taken:
                continue
            alias_slugs = {_slug(a) for a in aliases}
            if slug in alias_slugs:
                mapping[key] = idx
                taken.add(key)
                break
            # Loose fallbacks for numeric/score-ish columns
            if key == "cvss" and ("cvss" in slug or "base score" in slug):
                mapping[key] = idx
                taken.add(key)
                break
            if key == "cves" and "cve" in slug:
                mapping[key] = idx
                taken.add(key)
                break
    return mapping


def _cell(row: list, col_map: dict[str, int], key: str):
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    return str(value).strip()


def _raw_from_cells(row: list, col_map: dict[str, int], source_tool: str) -> dict:
    severity_cell = _cell(row, col_map, "severity")
    cvss = _float(_cell(row, col_map, "cvss"))

    # Qualys uses a 1-5 numeric scale — translate to Nessus-style 0-4 first.
    if source_tool == "qualys" and severity_cell and re.fullmatch(r"\d+", severity_cell):
        q = int(severity_cell)
        qualys_map = {5: 4, 4: 3, 3: 2, 2: 1, 1: 0}
        if q in qualys_map:
            severity_cell = str(qualys_map[q])

    label, rank = parse_severity(severity_cell, cvss)

    title = _cell(row, col_map, "title") or ""
    cves = [
        c.strip()
        for c in re.split(r"[,; ]+", _cell(row, col_map, "cves") or "")
        if c.strip() and c.strip().lower().startswith("cve-")
    ]
    refs = [
        r for r in re.split(r"[\s,;]+", _cell(row, col_map, "references") or "")
        if r and r.startswith(("http://", "https://", "www."))
    ]

    raw_status = (_cell(row, col_map, "status") or "pending").strip().lower()
    status_map = {
        "solve": "solved",
        "solved": "solved",
        "resolved": "solved",
        "ignore": "ignore",
        "false positive": "false_positive",
        "false-positive": "false_positive",
        "false_positive": "false_positive",
        "pending": "pending",
    }
    normalized_status = status_map.get(raw_status, raw_status.replace(" ", "_"))
    comment = _cell(row, col_map, "comment") or ""

    return {
        "host": _cell(row, col_map, "host") or _cell(row, col_map, "hostname") or "",
        "os": _cell(row, col_map, "operating_system") or _cell(row, col_map, "os") or "",
        "port": _int_or_none(_cell(row, col_map, "port")),
        "protocol": _cell(row, col_map, "protocol") or "",
        "service": _cell(row, col_map, "service") or "",
        "plugin_id": _cell(row, col_map, "plugin_id") or "",
        "plugin_name": title,
        "plugin_family": _cell(row, col_map, "plugin_family") or "",
        "title": title,
        "severity_label": label,
        "severity": rank,
        "cvss_score": cvss,
        "cvss_vector": _cell(row, col_map, "cvss_vector") or "",
        "description": _cell(row, col_map, "description") or "",
        "synopsis": "",
        "solution": _cell(row, col_map, "solution") or "",
        "references": refs,
        "cves": cves,
        "evidence": _cell(row, col_map, "evidence") or "",
        "status": normalized_status or "pending",
        "comment": comment,
        "mac_address": _cell(row, col_map, "mac_address") or "",
        "hostname": _cell(row, col_map, "hostname") or "",
        "operating_system": _cell(row, col_map, "operating_system") or _cell(row, col_map, "os") or "",
        "remarks": comment,
    }


# ─── CSV ──────────────────────────────────────────────────────────────────────

def parse_csv(content: bytes) -> tuple[list[dict], str]:
    """Parse a CSV export with flexible header detection."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty.")

    # Skip leading blank rows to find the header row.
    header_idx = next(
        (i for i, r in enumerate(rows) if any(c.strip() for c in r)),
        0,
    )
    headers = rows[header_idx]
    data_rows = rows[header_idx + 1:]

    source_tool = detect_source_tool(headers)
    col_map = _build_header_map(headers)
    if "title" not in col_map and "host" not in col_map:
        raise ValueError(
            "Could not recognize the CSV columns — expected a Nessus / OpenVAS / "
            "Qualys export (e.g. columns like 'Plugin Name', 'Host', 'Severity')."
        )

    raw = []
    for row in data_rows:
        if not any(c and c.strip() for c in row):
            continue
        raw.append(_raw_from_cells(row, col_map, source_tool))
    return raw, source_tool


# ─── Excel (.xls) ────────────────────────────────────────────────────────────

def parse_xls(content: bytes) -> tuple[list[dict], str]:
    """Parse an .xls export with flexible column detection."""
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("Excel (.xls) support requires 'xlrd'.") from exc

    try:
        wb = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ValueError(f"Could not read the .xls file: {exc}")

    try:
        sheet = wb.sheet_by_index(0)
        if sheet.nrows > 200_000:
            raise ValueError("Worksheet has too many rows to import (limit 200,000).")
        if sheet.ncols > 1_000:
            raise ValueError("Worksheet has too many columns to import (limit 1,000).")

        rows = []
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            rows.append(["" if v is None else str(v).strip() for v in row])

        header_idx = next(
            (i for i, r in enumerate(rows) if any(c for c in r)),
            0,
        )
        headers = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        source_tool = detect_source_tool(headers)
        col_map = _build_header_map(headers)
        if "title" not in col_map and "host" not in col_map:
            raise ValueError(
                "Could not recognize the worksheet columns — expected a Nessus / "
                "OpenVAS / Qualys export (e.g. columns like 'Plugin Name', 'Host', "
                "'Severity')."
            )

        raw = []
        for row in data_rows:
            if not any(c for c in row):
                continue
            raw.append(_raw_from_cells(row, col_map, source_tool))
        return raw, source_tool
    finally:
        wb.release_resources()


# ─── Excel (.xlsx) ────────────────────────────────────────────────────────────

def parse_xlsx(content: bytes) -> tuple[list[dict], str]:
    """Parse an .xlsx export with flexible column detection."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("Excel (.xlsx) support requires 'openpyxl'.") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the .xlsx file: {exc}")

    try:
        sheet = wb.active
        # Bounds guard against pathological / zip-bomb style workbooks.
        if sheet.max_row and sheet.max_row > 200_000:
            raise ValueError("Worksheet has too many rows to import (limit 200,000).")
        if sheet.max_column and sheet.max_column > 1_000:
            raise ValueError("Worksheet has too many columns to import (limit 1,000).")

        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v).strip() for v in row])

        header_idx = next(
            (i for i, r in enumerate(rows) if any(c for c in r)),
            0,
        )
        headers = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        source_tool = detect_source_tool(headers)
        col_map = _build_header_map(headers)
        if "title" not in col_map and "host" not in col_map:
            raise ValueError(
                "Could not recognize the worksheet columns — expected a Nessus / "
                "OpenVAS / Qualys export (e.g. columns like 'Plugin Name', 'Host', "
                "'Severity')."
            )

        raw = []
        for row in data_rows:
            if not any(c for c in row):
                continue
            raw.append(_raw_from_cells(row, col_map, source_tool))
        return raw, source_tool
    finally:
        wb.close()


# ─── Public entry point ───────────────────────────────────────────────────────

def parse_upload(content: bytes, filename: str) -> tuple[list[dict], str, str]:
    """Parse an uploaded export file.

    Returns ``(raw_findings, source_tool, file_format)``. Raises ``ValueError``
    for unsupported / unrecognizable files.
    """
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type '{ext}'. Upload a .nessus, .xml, .csv or "
            f".xlsx export (max {VAPT_MAX_FILE_SIZE_MB} MB)."
        )
    if len(content) > MAX_FILE_SIZE:
        raise ValueError(f"File exceeds the {VAPT_MAX_FILE_SIZE_MB} MB size limit.")

    if ext in (".nessus", ".xml"):
        raw = parse_nessus_xml(content)
        return raw, "nessus", "xml"

    if ext == ".csv":
        raw, source_tool = parse_csv(content)
        return raw, source_tool, "csv"

    if ext == ".xls":
        raw, source_tool = parse_xls(content)
        return raw, source_tool, "xls"

    raw, source_tool = parse_xlsx(content)
    return raw, source_tool, "xlsx"
